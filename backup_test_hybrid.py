#!/usr/bin/env python3
"""
Test hybrid approach: CellRichText -> xlsxwriter -> openpyxl roundtrip
This tests the exact scenario used in the main app.
"""

import logging
import sys
import openpyxl
import xlsxwriter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles.colors import Color

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

def create_test_cellrichtext():
    """Create a test CellRichText object similar to what the main app creates"""
    
    # Create parts similar to what the app generates
    parts = []
    
    # Plain text
    parts.append("Plain text at start.\n\n")
    
    # Bold text
    bold_font = InlineFont(b=True)
    parts.append(TextBlock(bold_font, "This is bold text. "))
    
    # Red text  
    red_font = InlineFont(color=Color(rgb="FF0000"))
    parts.append(TextBlock(red_font, "This is red text. "))
    
    # Blue italic text
    blue_italic_font = InlineFont(i=True, color=Color(rgb="0000FF"))
    parts.append(TextBlock(blue_italic_font, "This is blue italic text."))
    
    # More plain text
    parts.append("\n\nMore plain text at end.")
    
    return CellRichText(*parts)

def test_pure_openpyxl():
    """Test 1: Pure openpyxl approach (control test)"""
    logger.info("=== Test 1: Pure openpyxl approach ===")
    
    try:
        # Create test data
        rich_text = create_test_cellrichtext()
        logger.info(f"Original CellRichText has {len(rich_text)} parts")
        
        # Write with openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = rich_text
        wb.save("test_pure_openpyxl.xlsx")
        wb.close()
        
        # Read back with openpyxl
        wb_read = openpyxl.load_workbook("test_pure_openpyxl.xlsx", rich_text=True)
        ws_read = wb_read.active
        read_value = ws_read['A1'].value
        
        logger.info(f"Read back - Type: {type(read_value)}")
        logger.info(f"Read back - Parts: {len(read_value) if hasattr(read_value, '__len__') else 'N/A'}")
        
        # Compare
        original_text = str(rich_text)
        read_text = str(read_value)
        
        if original_text == read_text:
            logger.info("✓ Pure openpyxl: TEXT PRESERVED")
        else:
            logger.error("✗ Pure openpyxl: CORRUPTION DETECTED")
            
        wb_read.close()
        
    except Exception as e:
        logger.error(f"Pure openpyxl test failed: {e}")

def convert_cellrichtext_to_xlsxwriter_format(rich_text_obj, workbook):
    """Convert CellRichText to xlsxwriter rich string format (from main app)"""
    try:
        if not hasattr(rich_text_obj, '__iter__'):
            return str(rich_text_obj)
            
        logger.info(f"Converting CellRichText with {len(rich_text_obj)} parts")
        
        rich_parts = []
        for i, part in enumerate(rich_text_obj):
            logger.info(f"Processing part {i}: {type(part)}")
            
            if hasattr(part, 'text') and hasattr(part, 'font'):
                # TextBlock with formatting
                format_dict = {}
                if hasattr(part.font, 'b') and part.font.b:
                    format_dict['bold'] = True
                if hasattr(part.font, 'i') and part.font.i:
                    format_dict['italic'] = True
                if hasattr(part.font, 'color') and part.font.color:
                    if hasattr(part.font.color, 'rgb') and part.font.color.rgb:
                        color_rgb = part.font.color.rgb
                        # Convert to hex format for xlsxwriter
                        if len(color_rgb) == 8 and color_rgb.startswith('00'):
                            color_hex = '#' + color_rgb[2:]  # Remove alpha channel
                        else:
                            color_hex = '#' + color_rgb
                        format_dict['color'] = color_hex
                
                if format_dict:
                    format_obj = workbook.add_format(format_dict)
                    rich_parts.extend([format_obj, part.text])
                    logger.info(f"Added formatted part: {format_dict} - '{part.text[:30]}...'")
                else:
                    rich_parts.append(part.text)
                    logger.info(f"Added plain text from TextBlock: '{part.text[:30]}...'")
                    
            elif isinstance(part, str):
                # Plain string
                rich_parts.append(part)
                logger.info(f"Added string part: '{part[:30]}...'")
            else:
                # Other type
                part_str = str(part)
                rich_parts.append(part_str)
                logger.info(f"Added other type {type(part)}: '{part_str[:30]}...'")
        
        return rich_parts
        
    except Exception as e:
        logger.error(f"Error converting to xlsxwriter format: {e}")
        return str(rich_text_obj)

def test_hybrid_approach():
    """Test 2: Hybrid approach (CellRichText -> xlsxwriter -> openpyxl read)"""
    logger.info("=== Test 2: Hybrid approach (like main app) ===")
    
    try:
        # Create test data
        original_rich_text = create_test_cellrichtext()
        logger.info(f"Original CellRichText has {len(original_rich_text)} parts")
        
        # Write with xlsxwriter (simulating main app's approach)
        workbook = xlsxwriter.Workbook("test_hybrid.xlsx")
        worksheet = workbook.add_worksheet()
        
        # Convert to xlsxwriter format
        rich_parts = convert_cellrichtext_to_xlsxwriter_format(original_rich_text, workbook)
        
        if len(rich_parts) > 0:
            logger.info(f"xlsxwriter rich_parts count: {len(rich_parts)}")
            worksheet.write_rich_string(0, 0, *rich_parts)
        else:
            worksheet.write(0, 0, str(original_rich_text))
            
        workbook.close()
        
        # Read back with openpyxl
        wb_read = openpyxl.load_workbook("test_hybrid.xlsx", rich_text=True)
        ws_read = wb_read.active
        read_value = ws_read['A1'].value
        
        logger.info(f"Read back from xlsxwriter file - Type: {type(read_value)}")
        
        if hasattr(read_value, '__iter__') and hasattr(read_value, '__len__'):
            logger.info(f"Read back - Parts: {len(read_value)}")
            
            # Log each part in detail
            for i, part in enumerate(read_value):
                if hasattr(part, 'text'):
                    logger.info(f"  Part {i}: TextBlock '{part.text[:50]}...' (len={len(part.text)})")
                else:
                    logger.info(f"  Part {i}: String '{str(part)[:50]}...' (len={len(str(part))})")
        
        # Compare text content
        original_text = str(original_rich_text)
        read_text = str(read_value)
        
        logger.info(f"Original text: '{original_text}'")
        logger.info(f"Read text:     '{read_text}'")
        
        if original_text == read_text:
            logger.info("✓ Hybrid approach: TEXT PRESERVED")
        else:
            logger.error("✗ Hybrid approach: CORRUPTION DETECTED")
            logger.error("This is where the corruption happens!")
            
        wb_read.close()
        
    except Exception as e:
        logger.error(f"Hybrid test failed: {e}")

def test_xlsxwriter_only():
    """Test 3: xlsxwriter only (can't read back, but we can create and check)"""
    logger.info("=== Test 3: xlsxwriter only ===")
    
    try:
        # Create test data
        original_rich_text = create_test_cellrichtext()
        
        # Write with xlsxwriter
        workbook = xlsxwriter.Workbook("test_xlsxwriter_only.xlsx")
        worksheet = workbook.add_worksheet()
        
        # Convert to xlsxwriter format  
        rich_parts = convert_cellrichtext_to_xlsxwriter_format(original_rich_text, workbook)
        
        logger.info(f"xlsxwriter rich_parts: {len(rich_parts)} items")
        
        # Write rich string
        if len(rich_parts) > 0:
            worksheet.write_rich_string(0, 0, *rich_parts)
            logger.info("✓ xlsxwriter: Rich string written successfully")
        else:
            worksheet.write(0, 0, str(original_rich_text))
            logger.info("Fallback: Plain string written")
            
        workbook.close()
        
    except Exception as e:
        logger.error(f"xlsxwriter only test failed: {e}")

if __name__ == "__main__":
    logger.info("Testing rich text corruption in hybrid approach...")
    
    # Run all tests
    test_pure_openpyxl()
    test_hybrid_approach() 
    test_xlsxwriter_only()
    
    logger.info("All tests completed. Check the log output above for results.")