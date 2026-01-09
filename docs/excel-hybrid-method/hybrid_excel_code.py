"""
HYBRID EXCEL METHOD - Reference Documentation
==============================================

This file contains the extracted code for the hybrid Excel method (openpyxl + xlsxwriter)
that preserves text formatting when adding rows to Excel files.

This is reference documentation extracted from core/excel_manager.py.
For complete implementation, see the original file.

BREAKTHROUGH METHOD: This hybrid approach is the only working solution for maintaining
rich text, formulas, and formatting during Excel operations.

WHY HYBRID?
- openpyxl: Excellent for READING Excel files and preserving formulas/rich text
- xlsxwriter: Excellent for WRITING Excel files with precise formatting control
- Neither library alone can handle the full workflow

WORKFLOW STEPS:
1. Read existing Excel file with openpyxl (preserves all data and formatting)
2. Create new temporary file with xlsxwriter (allows precise formatting control)
3. Copy all existing data with formatting preserved
4. Add new row with special data handling
5. Atomically replace original file

File: docs/excel-hybrid-method/hybrid_excel_code.py
Original: core/excel_manager.py (lines 281-792)
"""

import logging
import os
from datetime import datetime
from typing import Dict

import openpyxl
import xlsxwriter
from openpyxl.utils import get_column_letter

# NOTE: This file is reference documentation only.
# In the actual implementation (core/excel_manager.py), this imports:
# from core.field_definitions import field_manager
# The field_manager handles field visibility and custom field names.

logger = logging.getLogger(__name__)


# ==============================================================================
# MAIN HYBRID METHOD
# ==============================================================================

def add_row_with_xlsxwriter(self, data: Dict[str, str], filename: str, row_color: str = "none") -> bool:
    """
    BREAKTHROUGH METHOD: Add new row using hybrid approach: openpyxl for reading, xlsxwriter for writing

    This method solves the fundamental problem that neither library alone can handle:
    - openpyxl can read and preserve existing formatting but has poor rich text writing
    - xlsxwriter has excellent writing capabilities but cannot read existing files

    The hybrid approach combines both libraries' strengths to achieve:
    - Preserve all existing formulas (=TEXT() date formulas)
    - Preserve all existing rich text formatting (bold, italic, colors)
    - Preserve all existing background colors
    - Add new rows with full formatting support
    - Handle CellRichText corruption from library interoperability

    Args:
        data: Dictionary mapping column names to values
        filename: PDF filename to add to Händelse and Källa fields
        row_color: Background color for new row ('yellow', 'green', 'blue', 'pink', 'gray', 'none')

    Returns:
        True if successful
        "file_locked" if Excel file is locked by another application
        False if error occurs

    Workflow:
        STEP 1: Read with openpyxl (preserves formulas and rich text)
        STEP 2: Write with xlsxwriter (allows rich text and formula writing)
        STEP 3: Copy all existing data with formatting
        STEP 4: Append new row with special data handling
        STEP 5: Atomic file replacement
    """
    try:
        if not self.excel_path or not os.path.exists(self.excel_path):
            logger.error(f"Excel file not found: {self.excel_path}")
            return False

        # Check if Excel file is locked by another application
        try:
            with open(self.excel_path, 'r+b'):
                pass
        except (OSError, PermissionError):
            # File is locked - return special error code for caller to handle
            logger.warning(f"Excel file is locked: {self.excel_path}")
            return "file_locked"

        # ======================================================================
        # STEP 1: Read existing data with openpyxl (preserves formulas)
        # ======================================================================
        read_workbook = openpyxl.load_workbook(self.excel_path, rich_text=True)
        read_worksheet = read_workbook.active

        # Get all existing data including formulas AND formatting
        existing_data = []
        existing_formats = []

        for row_idx, row in enumerate(read_worksheet.iter_rows()):
            row_data = []
            row_formats = []
            for col_idx, cell in enumerate(row):
                # Capture cell data
                if cell.data_type == 'f':  # Formula
                    logger.info(f"Found formula at ({row_idx}, {col_idx}): {cell.value}")
                    row_data.append(('formula', cell.value))
                elif hasattr(cell.value, '__class__') and cell.value.__class__.__name__ == 'CellRichText':
                    # Fix corrupted CellRichText objects from openpyxl reading xlsxwriter files
                    repaired_richtext = self._repair_corrupted_cellrichtext(cell.value)
                    row_data.append(('richtext', repaired_richtext))
                else:
                    row_data.append(('value', cell.value))

                # Capture cell formatting with safe color extraction
                def safe_color_extract(color_obj):
                    try:
                        if color_obj and hasattr(color_obj, 'rgb'):
                            return str(color_obj.rgb) if color_obj.rgb else None
                        return None
                    except (AttributeError, TypeError, ValueError) as e:
                        logger.warning(f"Could not extract color: {e}")
                        return None

                cell_format = {
                    'font_bold': cell.font.bold if cell.font.bold else False,
                    'font_italic': cell.font.italic if cell.font.italic else False,
                    'font_color': safe_color_extract(cell.font.color),
                    'font_size': cell.font.size if cell.font.size else None,
                    'fill_color': safe_color_extract(cell.fill.start_color),
                    'alignment_wrap': cell.alignment.wrap_text if cell.alignment.wrap_text else False,
                    'alignment_horizontal': cell.alignment.horizontal if cell.alignment.horizontal else None,
                    'alignment_vertical': cell.alignment.vertical if cell.alignment.vertical else None,
                }
                row_formats.append(cell_format)

            existing_data.append(row_data)
            existing_formats.append(row_formats)

        # Capture column widths
        column_widths = {}
        for col_letter, dimension in read_worksheet.column_dimensions.items():
            if dimension.width:
                column_widths[col_letter] = dimension.width

        # Capture row heights
        row_heights = {}
        for row_num, dimension in read_worksheet.row_dimensions.items():
            if dimension.height:
                row_heights[row_num] = dimension.height

        # ======================================================================
        # STEP 2: Create new file with xlsxwriter
        # ======================================================================
        temp_file = f"{self.excel_path}.tmp"
        write_workbook = xlsxwriter.Workbook(temp_file)
        write_worksheet = write_workbook.add_worksheet()

        # Set default wrap text formatting for entire worksheet (A1:T100)
        wrap_format = write_workbook.add_format({'text_wrap': True})
        write_worksheet.set_column('A:T', None, wrap_format)

        # Apply column widths with wrap text formatting preserved
        for col_letter, width in column_widths.items():
            col_idx = openpyxl.utils.column_index_from_string(col_letter) - 1
            write_worksheet.set_column(col_idx, col_idx, width, wrap_format)

        # Apply row heights
        for row_num, height in row_heights.items():
            write_worksheet.set_row(row_num - 1, height)

        # ======================================================================
        # STEP 3: Copy existing data with formatting to new workbook
        # ======================================================================
        for row_idx, (row_data, row_formats) in enumerate(zip(existing_data, existing_formats)):
            for col_idx, ((data_type, value), cell_format) in enumerate(zip(row_data, row_formats)):
                # Create xlsxwriter format from openpyxl format
                format_dict = {}
                if cell_format['font_bold']:
                    format_dict['bold'] = True
                if cell_format['font_italic']:
                    format_dict['italic'] = True
                if cell_format['font_color']:
                    # Convert color to valid xlsxwriter format
                    color_hex = self._convert_color_to_hex(cell_format['font_color'])
                    # Skip default black color to avoid overriding normal text
                    if color_hex and color_hex not in ['#000000', '#000']:
                        format_dict['color'] = color_hex
                if cell_format['font_size']:
                    format_dict['size'] = cell_format['font_size']
                if cell_format['fill_color']:
                    # Convert color to valid xlsxwriter format
                    color_hex = self._convert_color_to_hex(cell_format['fill_color'])
                    # Skip default white/transparent background to avoid overriding normal cells
                    if color_hex and color_hex not in ['#FFFFFF', '#FFF', '#000000', '#000']:
                        format_dict['bg_color'] = color_hex
                # Always enable text wrap to preserve user's worksheet setting
                format_dict['text_wrap'] = True
                if cell_format['alignment_horizontal']:
                    format_dict['align'] = cell_format['alignment_horizontal']
                if cell_format['alignment_vertical']:
                    format_dict['valign'] = cell_format['alignment_vertical']

                # Create format object - always include text_wrap to preserve worksheet setting
                cell_format_obj = write_workbook.add_format(format_dict) if format_dict else write_workbook.add_format({'text_wrap': True})

                # Write data with formatting
                if data_type == 'formula' and value:
                    logger.info(f"Writing formula at ({row_idx}, {col_idx}): {value}")
                    write_worksheet.write_formula(row_idx, col_idx, value, cell_format_obj)
                elif data_type == 'richtext':
                    # Convert openpyxl RichText to xlsxwriter rich string
                    # Extract row color from existing cell format to preserve background colors
                    detected_row_color = self._extract_row_color_from_format(cell_format)
                    logger.info(f"DEBUG: Detected row color '{detected_row_color}' for existing rich text at ({row_idx}, {col_idx})")
                    self._write_rich_text_xlsxwriter(write_worksheet, row_idx, col_idx, value, write_workbook, cell_format_obj, detected_row_color)
                elif value is not None:
                    write_worksheet.write(row_idx, col_idx, value, cell_format_obj)
                elif cell_format_obj:
                    # Apply formatting even to empty cells
                    write_worksheet.write_blank(row_idx, col_idx, None, cell_format_obj)

        # ======================================================================
        # STEP 4: Add new row with rich text support
        # ======================================================================
        next_row = read_worksheet.max_row  # Use actual last row with data
        special_data = self._prepare_special_data(data, filename)

        # Create a default format with word wrap for new cells
        # Include background color if specified
        if row_color and row_color != "none":
            color_map = {
                "yellow": "#FFFF99",
                "green": "#CCFFCC",
                "blue": "#CCE5FF",
                "pink": "#FFCCEE",
                "gray": "#E6E6E6"
            }
            format_dict = {'text_wrap': True}  # Always preserve text wrap
            if row_color in color_map:
                format_dict['bg_color'] = color_map[row_color]
            default_format = write_workbook.add_format(format_dict)
        else:
            default_format = write_workbook.add_format({'text_wrap': True})

        # Only write VISIBLE columns (respects field visibility settings)
        # NOTE: field_manager is imported from core.field_definitions in actual implementation
        # visible_column_names = field_manager.get_visible_display_names()
        # For reference documentation, assume all columns are visible:
        visible_column_names = list(self.columns.keys())
        for col_name, col_idx in self.columns.items():
            # Skip hidden columns
            if col_name not in visible_column_names:
                continue

            value = special_data.get(col_name, '')

            # Special handling for Dag column - create formula
            if col_name == 'Dag' and not value:
                # Create formula =TEXT(I{row};"ddd") where I is the Startdatum column
                tid_start_col_idx = self.columns.get('Startdatum', 9)  # Default to column I (9)
                formula = f'=TEXT({get_column_letter(tid_start_col_idx)}{next_row + 1},"ddd")'
                logger.info(f"Creating Dag formula for new row: {formula}")
                write_worksheet.write_formula(next_row, col_idx-1, formula, default_format)
            elif hasattr(value, '__class__') and value.__class__.__name__ == 'CellRichText':
                # Convert CellRichText to xlsxwriter rich string with background color
                self._write_rich_text_xlsxwriter(write_worksheet, next_row, col_idx-1, value, write_workbook, default_format, row_color)
            elif value is not None:
                write_worksheet.write(next_row, col_idx-1, value, default_format)
            else:
                # Apply word wrap format even to empty cells to maintain consistency
                write_worksheet.write_blank(next_row, col_idx-1, None, default_format)

        write_workbook.close()

        # ======================================================================
        # STEP 5: Replace original file atomically
        # ======================================================================
        if os.path.exists(self.excel_path):
            os.replace(temp_file, self.excel_path)

        logger.info(f"Added row to Excel file using xlsxwriter hybrid approach at row {next_row + 1}")
        return True

    except Exception as e:
        logger.error(f"Error adding row with xlsxwriter: {e}")
        # Clean up temp file if it exists
        if 'temp_file' in locals() and os.path.exists(temp_file):
            os.remove(temp_file)
        return False


# ==============================================================================
# HELPER METHODS
# ==============================================================================

def _prepare_special_data(self, data: Dict[str, str], filename: str) -> Dict[str, str]:
    """
    Prepare data with special column handling (extracted from original add_row)

    This method handles application-specific business logic for certain columns:
    - Händelse: Preserves user content, appends filename if needed
    - Startdatum: Uses date from filename if user hasn't filled it
    - Källa: Only uses generated filename if field is empty

    Args:
        data: Original data dictionary
        filename: PDF filename to add to specific fields

    Returns:
        Modified data dictionary with special column handling applied
    """
    special_data = data.copy()

    # Händelse - preserve user content and add filename if needed
    if 'Händelse' in self.columns:
        current_content = data.get('Händelse', '')
        if hasattr(current_content, '__class__') and current_content.__class__.__name__ == 'CellRichText':
            has_content = len(current_content) > 0
            if has_content and filename:
                special_data['Händelse'] = current_content
            elif not has_content:
                special_data['Händelse'] = filename if filename else ""
            else:
                special_data['Händelse'] = current_content
        else:
            current_content = str(current_content).strip()
            if current_content:
                if filename and filename not in current_content:
                    special_data['Händelse'] = f"{current_content}\n{filename}"
                else:
                    special_data['Händelse'] = current_content
            else:
                if filename:
                    special_data['Händelse'] = f"\n\n{filename}"
                else:
                    special_data['Händelse'] = ""

    # Startdatum - use date from filename if user hasn't filled it in
    if 'Startdatum' in self.columns and 'date' in data:
        user_tid_start = str(special_data.get('Startdatum', '')).strip()
        if not user_tid_start:
            try:
                date_obj = datetime.strptime(data['date'], '%Y-%m-%d')
                special_data['Startdatum'] = date_obj.date()
            except ValueError:
                special_data['Startdatum'] = data.get('date', '')

    # Källa - only use generated filename if field is empty AND we have a filename
    if 'Källa' in self.columns:
        current_kalla1 = data.get('Källa', '').strip()
        if not current_kalla1 and filename:
            special_data['Källa'] = filename
        else:
            special_data['Källa'] = current_kalla1

    return special_data


def _convert_color_to_hex(self, color_value):
    """
    Convert various color formats to xlsxwriter-compatible hex string

    openpyxl and xlsxwriter use different color representation formats.
    This method translates between them to ensure colors are preserved.

    Args:
        color_value: Color in openpyxl format (various representations possible)

    Returns:
        Hex color string like "#FFFF99" or None if conversion fails

    Handles:
        - ARGB format (8 hex chars) → RGB format (6 hex chars)
        - RGB format (6 hex chars) → Hex string with #
        - Color names → Hex equivalents
        - Invalid/default colors → None (to avoid overriding defaults)
    """
    try:
        if not color_value:
            return None

        # Convert to string first
        color_str = str(color_value)

        # Remove any non-hex characters and get just the hex part
        import re
        hex_match = re.search(r'([0-9A-Fa-f]{6,8})', color_str)
        if hex_match:
            hex_color = hex_match.group(1)
            # Take last 6 characters (RGB) if 8 characters (ARGB)
            if len(hex_color) == 8:
                hex_color = hex_color[2:]  # Remove alpha channel
            return f"#{hex_color}"

        # If no hex found, try some common color names
        color_lower = color_str.lower()
        color_map = {
            'black': '#000000',
            'white': '#FFFFFF',
            'red': '#FF0000',
            'green': '#00FF00',
            'blue': '#0000FF',
        }
        return color_map.get(color_lower, None)

    except Exception as e:
        logger.warning(f"Could not convert color {color_value}: {e}")
        return None


def _extract_row_color_from_format(self, cell_format: Dict) -> str:
    """
    Extract row color name from cell format background color

    When copying existing rows, we need to detect what background color
    they had so we can preserve it. This method reverse-maps from hex
    colors back to semantic color names.

    Args:
        cell_format: Dictionary with cell formatting properties

    Returns:
        Row color name ('yellow', 'green', 'blue', 'pink', 'gray') or None
    """
    fill_color = cell_format.get('fill_color')
    if not fill_color:
        return None

    # Convert to hex format for comparison
    color_hex = self._convert_color_to_hex(fill_color)
    if not color_hex:
        return None

    # Map hex colors back to our row color names
    color_reverse_map = {
        "#FFFF99": "yellow",
        "#CCFFCC": "green",
        "#CCE5FF": "blue",
        "#FFCCEE": "pink",
        "#E6E6E6": "gray"
    }

    return color_reverse_map.get(color_hex.upper())


def _write_rich_text_xlsxwriter(self, worksheet, row, col, rich_text_obj, workbook, base_format=None, row_color=None):
    """
    BREAKTHROUGH METHOD: Convert openpyxl CellRichText to xlsxwriter rich string

    This is one of the most complex parts of the hybrid method. It handles the
    conversion of rich text (text with mixed formatting like bold, colors, italic)
    from openpyxl's representation to xlsxwriter's representation.

    Critical Edge Cases Handled:
    1. Uniform formatting (entire text same color/style) - xlsxwriter API limitation
    2. Mixed formatting (different colors/styles) - standard rich text case
    3. Background colors on rich text cells - requires special API usage
    4. Plain text fallback - graceful degradation

    Args:
        worksheet: xlsxwriter worksheet object
        row: Row index (0-based)
        col: Column index (0-based)
        rich_text_obj: openpyxl CellRichText object to convert
        workbook: xlsxwriter workbook object (for creating formats)
        base_format: Base format to apply (includes background color)
        row_color: Background color name for the cell ('yellow', 'green', etc.)
    """
    logger.info(f"DEBUG: _write_rich_text_xlsxwriter called with row_color='{row_color}', base_format={base_format is not None}")
    try:
        # Extract base format properties (like background color and text wrap)
        base_format_dict = {}
        if base_format:
            # For xlsxwriter Format objects, we need to reconstruct the format dict
            # Since we know we're passing our own default_format, extract from the color mapping
            # This is safer than trying to access internal xlsxwriter format properties
            if row_color and row_color != "none":
                color_map = {
                    "yellow": "#FFFF99",
                    "green": "#CCFFCC",
                    "blue": "#CCE5FF",
                    "pink": "#FFCCEE",
                    "gray": "#E6E6E6"
                }
                if row_color in color_map:
                    base_format_dict['bg_color'] = color_map[row_color]
                    logger.info(f"DEBUG: Added bg_color '{color_map[row_color]}' to base_format_dict for row_color '{row_color}'")
                else:
                    logger.warning(f"DEBUG: row_color '{row_color}' not found in color_map")
            else:
                logger.info(f"DEBUG: No background color applied - row_color='{row_color}'")
            base_format_dict['text_wrap'] = True  # Always include text wrap
            logger.info(f"DEBUG: Final base_format_dict: {base_format_dict}")

        if not hasattr(rich_text_obj, '__iter__'):
            # Plain text - apply base format
            if base_format:
                worksheet.write(row, col, str(rich_text_obj), base_format)
            else:
                worksheet.write(row, col, str(rich_text_obj))
            return

        # Build rich string for xlsxwriter
        rich_parts = []
        for part in rich_text_obj:
            if hasattr(part, 'text') and hasattr(part, 'font'):
                # TextBlock with formatting - include base format properties
                format_dict = base_format_dict.copy()  # Start with base format
                if hasattr(part.font, 'b') and part.font.b:
                    format_dict['bold'] = True
                if hasattr(part.font, 'i') and part.font.i:
                    format_dict['italic'] = True
                if hasattr(part.font, 'color') and part.font.color:
                    # Convert color to xlsxwriter format using the helper function
                    color_hex = self._convert_color_to_hex(part.font.color.rgb)
                    if color_hex and color_hex not in ['#000000', '#000']:
                        format_dict['color'] = color_hex

                if format_dict:
                    format_obj = workbook.add_format(format_dict)
                    logger.info(f"DEBUG: Created format object with dict: {format_dict}")
                    rich_parts.extend([format_obj, part.text])
                else:
                    logger.info(f"DEBUG: No format applied to text part: '{part.text[:20]}...'")
                    rich_parts.append(part.text)
            elif isinstance(part, str):
                # Plain text string part
                rich_parts.append(part)
            else:
                # Other type - convert to string
                part_str = str(part)
                rich_parts.append(part_str)

        # Handle rich text writing with background color support
        if rich_parts:
            logger.info(f"DEBUG: Writing rich string with {len(rich_parts)} parts to cell ({row}, {col})")
            logger.info(f"DEBUG: Rich parts structure: {[type(p).__name__ for p in rich_parts]}")

            # ======================================================================
            # UNIFORM FORMATTING FIX (v1.17.16)
            # ======================================================================
            # xlsxwriter write_rich_string() is designed for mixed formatting and fails with uniform formatting
            # Pattern: [format_obj, "entire text content"] - xlsxwriter edge case
            if (len(rich_parts) == 2 and
                hasattr(rich_parts[0], '__class__') and 'Format' in str(type(rich_parts[0])) and
                isinstance(rich_parts[1], str)):

                logger.info("DEBUG: Detected uniform formatting - using write() instead of write_rich_string()")

                # Extract the format and text
                format_obj = rich_parts[0]
                text_content = rich_parts[1]

                # For uniform formatting, we need to recreate the format to include background color
                # Since xlsxwriter doesn't allow extracting format properties, we need to rebuild
                if base_format_dict.get('bg_color'):
                    # We have background color - need to combine with text formatting
                    # The format_obj contains the text formatting, but we need to add background
                    # Unfortunately, we can't extract from format_obj, so we'll use write() with text format
                    # and accept that background color might not work perfectly for uniform formatting
                    logger.info("DEBUG: Uniform formatting with background - text formatting takes priority")
                    worksheet.write(row, col, text_content, format_obj)
                else:
                    # No background color - simple uniform formatting works perfectly
                    worksheet.write(row, col, text_content, format_obj)
                    logger.info("DEBUG: Applied uniform formatting without background")

                return  # Exit early - uniform formatting handled

            # ======================================================================
            # MIXED FORMATTING - Standard rich text case
            # ======================================================================
            logger.info("DEBUG: Using write_rich_string() for mixed formatting")

            # Apply background color using correct xlsxwriter API
            if base_format_dict.get('bg_color'):
                try:
                    logger.info("DEBUG: Rich text with background color - using correct xlsxwriter API")

                    # Create a base format with background color for the entire cell
                    cell_bg_format = workbook.add_format({
                        'bg_color': base_format_dict['bg_color'],
                        'text_wrap': True
                    })

                    # Use correct xlsxwriter API: pass cell format as last parameter
                    logger.info("DEBUG: Writing rich text with background format as parameter")
                    worksheet.write_rich_string(row, col, *rich_parts, cell_bg_format)

                except Exception as e:
                    logger.warning(f"DEBUG: Background approach failed: {e}")
                    # Fallback to normal rich text without background
                    worksheet.write_rich_string(row, col, *rich_parts)
            else:
                # No background color - use normal rich text
                worksheet.write_rich_string(row, col, *rich_parts)
        else:
            # Plain text handling
            logger.info(f"DEBUG: Writing plain text to cell ({row}, {col}): '{str(rich_text_obj)[:30]}...'")
            if base_format:
                worksheet.write(row, col, str(rich_text_obj), base_format)
            else:
                worksheet.write(row, col, str(rich_text_obj))

    except Exception as e:
        logger.warning(f"Error converting rich text to xlsxwriter format: {e}")
        # Fallback to plain text
        worksheet.write(row, col, str(rich_text_obj))


def _repair_corrupted_cellrichtext(self, rich_text_obj):
    """
    Repair corrupted CellRichText objects that result from openpyxl reading xlsxwriter files.

    INTEROPERABILITY ISSUE:
    When xlsxwriter writes rich text and openpyxl reads it back, we get corrupted structure:

    Expected Structure:
        Part 0: TextBlock('Formatted part 1')
        Part 1: TextBlock('Formatted part 2')

    Actual Structure (corrupted):
        Part 0: str('WHOLE TEXT CONTENT...')  # Duplicated full text
        Part 1: TextBlock('Formatted part 1')  # Individual formatted parts
        Part 2: TextBlock('Formatted part 2')

    This corruption causes text duplication in Excel cells. This method detects
    and repairs the corruption by removing the duplicated first part.

    Detection Heuristic:
        - First part is a plain string (not TextBlock)
        - First part length > 70% of total TextBlock content
        - At least one TextBlock exists in remaining parts

    Args:
        rich_text_obj: Potentially corrupted CellRichText object

    Returns:
        Repaired CellRichText object or original if not corrupted/repair fails
    """
    try:
        if not hasattr(rich_text_obj, '__iter__') or len(rich_text_obj) <= 1:
            return rich_text_obj

        # Check if first part is a plain string containing most/all of the text
        first_part = rich_text_obj[0]
        if not isinstance(first_part, str):
            return rich_text_obj

        # Get total length of TextBlock parts
        textblock_length = 0
        textblock_count = 0
        for part in rich_text_obj[1:]:
            if hasattr(part, 'text') and hasattr(part, 'font'):  # TextBlock
                textblock_length += len(part.text)
                textblock_count += 1
            elif isinstance(part, str):
                textblock_length += len(part)

        # If first part is significantly longer than sum of other parts,
        # and we have TextBlocks, this is likely a corruption
        first_part_len = len(first_part)
        if (first_part_len > textblock_length * 0.7 and  # First part contains 70%+ of text
            textblock_count > 0):  # We have actual TextBlocks

            logger.info(f"REPAIR: Detected corrupted CellRichText - first part ({first_part_len} chars) vs TextBlocks ({textblock_length} chars)")

            # Create new CellRichText without the duplicated first part
            from openpyxl.cell.rich_text import CellRichText
            repaired_parts = list(rich_text_obj[1:])  # Skip first duplicated part

            if repaired_parts:
                repaired = CellRichText(*repaired_parts)
                logger.info(f"REPAIR: Created repaired CellRichText with {len(repaired)} parts")
                return repaired
            else:
                # Fallback to original if repair fails
                logger.warning("REPAIR: No parts left after repair, using original")
                return rich_text_obj
        else:
            # Not corrupted, return as-is
            return rich_text_obj

    except Exception as e:
        logger.warning(f"Error repairing CellRichText: {e}")
        return rich_text_obj
