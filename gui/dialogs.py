"""
Dialog management for DJs Timeline-maskin
Contains all dialog classes and methods extracted from main_window.py
"""

# Standard library imports
import logging
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox

# Third-party GUI imports
import customtkinter as ctk

# Local imports
from gui.utils import ToolTip

logger = logging.getLogger(__name__)


class DialogManager:
    """Manages all dialog operations for the main application"""

    def __init__(self, parent_app):
        """Initialize dialog manager with reference to parent application"""
        self.parent = parent_app

    def show_excel_help(self):
        """Show help dialog for Excel field configuration"""
        help_win = ctk.CTkToplevel()
        help_win.title("Excel-fält information")
        help_win.geometry("500x400")
        help_win.transient(self.parent.root)
        help_win.grab_set()

        # Center dialog
        help_win.update_idletasks()
        x = (help_win.winfo_screenwidth() // 2) - (500 // 2)
        y = (help_win.winfo_screenheight() // 2) - (400 // 2)
        help_win.geometry(f"500x400+{x}+{y}")

        # Main frame
        main_frame = ctk.CTkFrame(help_win)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Header
        ctk.CTkLabel(main_frame, text="Flexibel Excel-hantering",
                font=('Arial', 14, 'bold')).pack(pady=(0, 15))

        # Information text
        info_text = """DJs Timeline-maskin har nu ett flexibelt system för Excel-fält!

ANPASSA DINA EXCEL-FÄLT:
• Du kan definiera dina egna kolumnnamn via menyn "Verktyg → Konfigurera fält..."
• Välj vilka fält som ska vara synliga eller dolda
• Ändra visningsnamn för att matcha dina Excel-kolumner
• Spara dina inställningar som mallar för återanvändning

SKAPA EXCEL-DOKUMENT:
• Klicka på "Skapa Excel" för att generera en Excel-fil
• Filen skapas automatiskt med dina definierade kolumnnamn
• Endast synliga fält inkluderas i Excel-filen
• Perfekt formaterad och redo att användas

TIPS:
• Använd standardmallen som utgångspunkt
• Skapa olika mallar för olika projekt
• Dolda fält exkluderas automatiskt från Excel-operationer"""

        # Scrollable text area
        text_frame = ctk.CTkFrame(main_frame)
        text_frame.pack(fill="both", expand=True, pady=(0, 15))

        text_area = ctk.CTkTextbox(text_frame, wrap="word",
                                   font=ctk.CTkFont(size=12), height=250)
        text_area.pack(fill="both", expand=True)
        text_area.insert("1.0", info_text)
        text_area.configure(state="disabled")

        # Buttons frame
        buttons_frame = ctk.CTkFrame(main_frame)
        buttons_frame.pack(fill="x", pady=(10, 0))

        # Configure fields button
        config_btn = ctk.CTkButton(buttons_frame, text="Konfigurera fält",
                               command=lambda: [help_win.destroy(), self.parent._show_field_config_dialog()],
                               width=120)
        config_btn.pack(side="left", padx=(0, 10))
        ToolTip(config_btn, "Öppna fältkonfigurationsdialogen för att anpassa dina Excel-kolumner")

        # Create template button
        template_btn = ctk.CTkButton(buttons_frame, text="Skapa Excel",
                               command=lambda: self.create_excel_template(help_win),
                               width=100)
        template_btn.pack(side="left", padx=(0, 10))
        ToolTip(template_btn, "Skapa en ny Excel-fil med dina definierade kolumner")

        # Close button
        close_btn = ctk.CTkButton(buttons_frame, text="Stäng",
                            command=help_win.destroy,
                            width=80)
        close_btn.pack(side="right")

    def create_excel_template(self, parent_window=None):
        """Create a new Excel file with the correct column headers"""
        try:
            # Ask user where to save the template
            template_path = filedialog.asksaveasfilename(
                title="Spara Excel-mall som...",
                defaultextension=".xlsx",
                filetypes=[("Excel-filer", "*.xlsx")],
                initialfile="Timeline_mall.xlsx"
            )

            if not template_path:
                return

            # Create new workbook
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Timeline"

            # Get current field display names (may be custom names) - only visible fields
            from core.field_definitions import field_manager
            headers = field_manager.get_visible_display_names()

            # Add headers to first row
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)

            # Style the header row and set up column formatting
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill

                # Set up column-specific formatting for the entire column
                column_letter = cell.column_letter

                if header == 'OBS':
                    # Text format for OBS column - ensure it's formatted as text
                    pass  # Formatting will be applied when data is added
                elif header in ['Startdatum', 'Slutdatum']:
                    # Date format for date columns - don't pre-format empty rows
                    pass  # Formatting will be applied when data is added
                elif header == 'Dag':
                    # Day format for Dag column - formula will be added when data is added
                    pass  # Formatting will be applied when data is added
                elif header.startswith('Note') or header == 'Händelse':
                    # Text wrapping for text fields - don't pre-format empty rows
                    pass  # Formatting will be applied when data is added

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (AttributeError, TypeError) as e:
                        logger.debug(f"Could not process cell value for width: {e}")
                        pass
                adjusted_width = min(max_length + 2, 50)  # Max width 50
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook
            wb.save(template_path)

            # Success message with option to open the created file
            result = messagebox.askyesno(
                "Mall skapad",
                f"Excel-mallen har skapats:\n{Path(template_path).name}\n\n" +
                "Vill du öppna mallen direkt i applikationen?"
            )

            if result:
                # Close help window if it exists
                if parent_window:
                    parent_window.destroy()

                # Load the created template
                if self.parent.excel_manager.load_excel_file(template_path):
                    self.parent.excel_path_var.set(Path(template_path).name)
                    # Save Excel file path to config for persistence
                    self.parent.config['excel_file'] = template_path
                    self.parent.config_manager.save_config(self.parent.config)
                    # No need to create fields - they're already created in setup_gui
                    # Enable the "Open Excel" button for newly created template
                    self.parent.open_excel_btn.configure(state="normal")
                    logger.info(f"Loaded created template: {template_path}")

        except Exception as e:
            messagebox.showerror("Fel", f"Kunde inte skapa Excel-mall: {str(e)}")
            logger.error(f"Error creating Excel template: {e}")

    def handle_paste_event(self, event, column_name):
        """Handle paste events with length checking and smart splitting"""
        try:
            # Get the text widget
            text_widget = event.widget

            # Create undo separator BEFORE any paste operation for Text widgets
            if isinstance(text_widget, tk.Text):
                try:
                    text_widget.edit_separator()
                    logger.info(f"Added undo separator before paste in {column_name}")
                except tk.TclError:
                    pass

            # Get clipboard content
            clipboard_content = self.parent.root.clipboard_get()

            # Check if clipboard content exceeds limit
            limit = self.parent.handelse_char_limit if column_name == 'Händelse' else self.parent.char_limit
            if len(clipboard_content) <= limit:
                # Normal paste - let it proceed but ensure undo separator was added
                logger.info(f"Normal paste in {column_name}: {len(clipboard_content)} chars")
                return False  # Don't block the event

            # Content is too long - offer options
            dialog_win = ctk.CTkToplevel()
            dialog_win.title("Text för lång")
            dialog_win.geometry("650x325")
            dialog_win.transient(self.parent.root)
            dialog_win.grab_set()

            # Center dialog
            dialog_win.update_idletasks()
            x = (dialog_win.winfo_screenwidth() // 2) - (650 // 2)
            y = (dialog_win.winfo_screenheight() // 2) - (325 // 2)
            dialog_win.geometry(f"650x325+{x}+{y}")

            # Dialog result variable
            dialog_result = [None]  # Use list to allow modification in nested functions

            # Main frame
            main_frame = ctk.CTkFrame(dialog_win)
            main_frame.pack(fill="both", expand=True, padx=20, pady=20)

            # Message
            message_text = (f"Den inklistrade texten är {len(clipboard_content)} tecken lång, "
                          f"vilket överstiger gränsen på {limit} tecken.\n\n"
                          f"Vad vill du göra?")
            ctk.CTkLabel(main_frame, text=message_text, font=ctk.CTkFont(size=10),
                    wraplength=580, justify="left").pack(pady=(0, 20))

            # Button frame
            button_frame = ctk.CTkFrame(main_frame)
            button_frame.pack(fill="x", pady=(10, 0))

            def on_truncate():
                dialog_result[0] = 'truncate'
                dialog_win.destroy()

            def on_split():
                dialog_result[0] = 'split'
                dialog_win.destroy()

            def on_cancel():
                dialog_result[0] = 'cancel'
                dialog_win.destroy()

            # Buttons with clear labels
            ctk.CTkButton(button_frame, text=f"Klipp av texten (första {limit} tecken)",
                     command=on_truncate, width=35).pack(pady=(0, 5))

            ctk.CTkButton(button_frame, text="Dela upp på flera fält",
                     command=on_split, fg_color="#17a2b8", width=35).pack(pady=(0, 5))

            ctk.CTkButton(button_frame, text="Avbryt inklistring",
                     command=on_cancel, width=150).pack(pady=(0, 5))

            # Wait for dialog to close
            dialog_win.wait_window()
            result = dialog_result[0]

            if result == 'cancel':  # Cancel
                return True  # Block the paste
            elif result == 'truncate':  # Truncate - paste first characters up to limit
                truncated_content = clipboard_content[:limit]
                text_widget = event.widget

                # Add undo separator before making changes
                if isinstance(text_widget, tk.Text):
                    try:
                        text_widget.edit_separator()
                    except tk.TclError:
                        pass

                text_widget.delete("1.0", tk.END)
                text_widget.insert("1.0", truncated_content)

                # Add undo separator after making changes
                if isinstance(text_widget, tk.Text):
                    try:
                        text_widget.edit_separator()
                    except tk.TclError:
                        pass

                self.parent.check_character_count(event, column_name)
                return True  # Block the original paste
            elif result == 'split':  # Split - try to split across fields
                return self.handle_text_splitting(clipboard_content, column_name)
            else:  # No dialog result (dialog was closed)
                return True  # Block the paste

        except tk.TclError:
            # No clipboard content
            return False

    def handle_text_splitting(self, text_content, start_column):
        """Handle splitting long text across multiple text fields"""
        # Define the text fields in order for splitting
        text_fields_order = ['Händelse', 'Note1', 'Note2', 'Note3']

        # Find starting position
        try:
            start_idx = text_fields_order.index(start_column)
        except ValueError:
            messagebox.showerror("Fel", "Texdelning stöds endast för Händelse, Note1, Note2 och Note3")
            return True  # Block paste

        # Get available fields from start position onwards
        available_fields = text_fields_order[start_idx:]

        # Check if any target fields have content and warn user
        fields_with_content = []
        for field_name in available_fields:
            if field_name in self.parent.excel_vars:
                widget = self.parent.excel_vars[field_name]
                if hasattr(widget, 'get'):
                    content = widget.get("1.0", tk.END).strip()
                    if content:
                        fields_with_content.append(field_name)

        # Warn about overwriting existing content
        if fields_with_content:
            overwrite_warning = "Följande fält innehåller redan text som kommer att skrivas över:\n• " + "\n• ".join(fields_with_content)
            confirm_overwrite = messagebox.askyesno(
                "Skriva över befintlig text?",
                f"{overwrite_warning}\n\nVill du fortsätta med texdelningen?"
            )
            if not confirm_overwrite:
                return True  # Block paste

        # Split text into chunks
        chunks = []
        remaining_text = text_content

        # Debug logging
        logger.info(f"Starting text splitting with {len(remaining_text)} characters")
        logger.info(f"First 50 chars: '{remaining_text[:50]}'")

        for field_name in available_fields:
            if len(remaining_text) == 0:
                break

            field_limit = self.parent.handelse_char_limit if field_name == 'Händelse' else self.parent.char_limit
            if len(remaining_text) <= field_limit:
                # Remaining text fits in this field
                chunks.append((field_name, remaining_text))
                logger.info(f"Final chunk for {field_name}: {len(remaining_text)} chars")
                remaining_text = ""  # Clear remaining text
                break
            else:
                # Find a good break point (try to break at word boundary)
                # Try to break at last space, newline, or punctuation within last 100 chars
                break_point = field_limit
                for i in range(min(100, field_limit)):
                    char_idx = field_limit - 1 - i
                    if char_idx < 0:
                        break
                    if char_idx < len(remaining_text):
                        char = remaining_text[char_idx]
                        if char in [' ', '\n', '.', '!', '?', ';', ':']:
                            # For punctuation, include it in current chunk
                            # For space/newline, don't include it in current chunk
                            if char in [' ', '\n']:
                                break_point = char_idx  # Don't include the space/newline
                            else:
                                break_point = char_idx + 1  # Include the punctuation
                            break

                chunk = remaining_text[:break_point].rstrip()  # Remove trailing whitespace
                chunks.append((field_name, chunk))

                # Debug logging
                logger.info(f"Chunk for {field_name}: {len(chunk)} chars, break_point: {break_point}")
                logger.info(f"Chunk ends with: '{chunk[-20:]}'")

                # Calculate actual chunk length after rstrip to avoid losing characters
                actual_chunk_length = len(chunk)
                remaining_text = remaining_text[actual_chunk_length:].lstrip() # Use actual chunk length, not break_point

                # More debug logging
                logger.info(f"Actual chunk length after rstrip: {actual_chunk_length}")
                logger.info(f"Remaining text starts with: '{remaining_text[:20]}'")
                logger.info(f"Remaining text length: {len(remaining_text)}")

        # Log final chunks summary
        for i, (field_name, chunk) in enumerate(chunks):
            logger.info(f"Final chunk {i+1} ({field_name}): {len(chunk)} chars, starts with: '{chunk[:20]}', ends with: '{chunk[-20:]}'")

        # Only show warning if text actually won't fit
        if remaining_text:
            # Create custom warning dialog
            warning_win = ctk.CTkToplevel()
            warning_win.title("Text för lång")
            warning_win.geometry("650x200")
            warning_win.transient(self.parent.root)
            warning_win.grab_set()

            # Center dialog
            warning_win.update_idletasks()
            x = (warning_win.winfo_screenwidth() // 2) - (650 // 2)
            y = (warning_win.winfo_screenheight() // 2) - (200 // 2)
            warning_win.geometry(f"650x200+{x}+{y}")

            # Main frame
            main_frame = ctk.CTkFrame(warning_win)
            main_frame.pack(fill="both", expand=True, padx=20, pady=20)

            # Warning message
            warning_text = ("Texten är för lång för att passa i tillgängliga fält. " +
                          f"Cirka {len(remaining_text)} tecken kommer att klippas bort från slutet.")
            ctk.CTkLabel(main_frame, text=warning_text, font=ctk.CTkFont(size=10),
                    wraplength=580, justify="left").pack(pady=(0, 20))

            # OK button
            ctk.CTkButton(main_frame, text="OK",
                     command=warning_win.destroy,
                     width=100).pack()

            # Wait for dialog to close
            warning_win.wait_window()

        # Show preview of how text will be split with meaningful excerpts
        preview_text = "Texten kommer att delas upp så här:\n\n"
        for field_name, chunk in chunks:
            # Show first and last few words to give better context
            words = chunk.split()
            if len(words) <= 10:
                preview = chunk
            else:
                first_words = ' '.join(words[:5])
                last_words = ' '.join(words[-5:])
                preview = f"{first_words} ... {last_words}"
            preview_text += f"• {field_name}: \"{preview}\" ({len(chunk)} tecken)\n"

        # Create custom dialog for split confirmation
        dialog_win = ctk.CTkToplevel()
        dialog_win.title("Bekräfta textuppdelning")
        dialog_win.geometry("650x400")
        dialog_win.transient(self.parent.root)
        dialog_win.grab_set()

        # Center dialog
        dialog_win.update_idletasks()
        x = (dialog_win.winfo_screenwidth() // 2) - (650 // 2)
        y = (dialog_win.winfo_screenheight() // 2) - (400 // 2)
        dialog_win.geometry(f"650x400+{x}+{y}")

        # Dialog result variable
        confirm_result = [False]

        # Main frame
        main_frame = ctk.CTkFrame(dialog_win)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Message with scrollable text area
        text_frame = ctk.CTkFrame(main_frame)
        text_frame.pack(fill="x", pady=(0, 15))

        text_area = ctk.CTkTextbox(text_frame, wrap="word",
                                   font=ctk.CTkFont(size=12), height=200, width=500)
        text_area.pack(fill="both")
        text_area.insert("1.0", preview_text + "\nFortsätt med denna uppdelning?")
        text_area.configure(state="disabled")

        # Button frame
        button_frame = ctk.CTkFrame(main_frame)
        button_frame.pack(fill="x", pady=(10, 0))

        def on_yes():
            confirm_result[0] = True
            dialog_win.destroy()

        def on_no():
            confirm_result[0] = False
            dialog_win.destroy()

        ctk.CTkButton(button_frame, text="Ja, fortsätt med uppdelning",
                 command=on_yes, fg_color="#28a745", width=25).pack(side="left", padx=(0, 10))

        ctk.CTkButton(button_frame, text="Nej, avbryt",
                 command=on_no, width=120).pack(side="left")

        # Wait for dialog to close
        dialog_win.wait_window()
        confirm_split = confirm_result[0]

        if confirm_split:
            # Apply the split text to fields
            for field_name, chunk in chunks:
                if field_name in self.parent.excel_vars:
                    widget = self.parent.excel_vars[field_name]
                    if hasattr(widget, 'delete'):
                        # Add undo separator before making changes (for Text widgets)
                        if isinstance(widget, tk.Text):
                            try:
                                widget.edit_separator()
                            except tk.TclError:
                                pass

                        # Temporarily unbind paste events to prevent conflicts
                        old_ctrl_v_binding = widget.bind('<Control-v>')
                        old_paste_binding = widget.bind('<<Paste>>')
                        widget.unbind('<Control-v>')
                        widget.unbind('<<Paste>>')

                        # Force widget to update and clear any pending events
                        widget.delete("1.0", tk.END)
                        widget.update_idletasks()  # Process any pending GUI events

                        # Insert the chunk
                        widget.insert("1.0", chunk)
                        widget.update_idletasks()  # Process insertion

                        # Add undo separator after making changes (for Text widgets)
                        if isinstance(widget, tk.Text):
                            try:
                                widget.edit_separator()
                            except tk.TclError:
                                pass

                        # Restore paste event bindings
                        if old_ctrl_v_binding:
                            widget.bind('<Control-v>', lambda e, col=field_name: self.parent.dialog_manager.handle_paste_event(e, col))
                        if old_paste_binding:
                            widget.bind('<<Paste>>', lambda e, col=field_name: self.parent.dialog_manager.handle_paste_event(e, col))

                        # Debug logging to verify what was actually inserted
                        actual_content = widget.get("1.0", tk.END).strip()
                        logger.info(f"Inserted into {field_name}: {len(actual_content)} chars")
                        logger.info(f"Actual content starts with: '{actual_content[:20]}'")
                        logger.info(f"Actual content ends with: '{actual_content[-20:]}'")

                        # Update character counter
                        fake_event = type('FakeEvent', (), {'widget': widget})()
                        self.parent.check_character_count(fake_event, field_name)

                        # Schedule a delayed verification to catch any interference
                        self.parent.root.after(100, lambda w=widget, fn=field_name, c=chunk: self.parent.verify_insertion(w, fn, c))

        return True  # Block the original paste
