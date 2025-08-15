#!/usr/bin/env python3
"""
BACKUP: Final Working Excel Hybrid Routine - v1.7.4
=====================================================

This file contains the COMPLETE WORKING Excel hybrid methods from v1.7.4.
These methods have been TESTED AND CONFIRMED to work perfectly for rich text Excel integration.

Date: 2025-07-22
Status: PRODUCTION READY - Successfully tested
Test Results: Perfect Excel rich text formatting with colors, bold, italic, line breaks

CRITICAL: These methods form the breakthrough hybrid approach:
- openpyxl for reading existing Excel data (preserves formulas)  
- xlsxwriter for writing new data (handles rich text perfectly)
- Method 2 character-by-character algorithm for text extraction
"""

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import openpyxl
import xlsxwriter
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelHybridMethods:
    """
    WORKING Excel Hybrid Methods - v1.7.4 PRODUCTION
    Contains all the breakthrough methods that make Excel rich text work perfectly.
    """
    
    def add_row_with_xlsxwriter(self, data: Dict[str, str], filename: str, row_color: str = "none") -> bool:
        """
        BREAKTHROUGH METHOD: Add new row using hybrid approach
        
        TESTED CONFIRMED WORKING:
        ✅ Rich text formatting preserved perfectly
        ✅ Colors, bold, italic all work correctly  
        ✅ Line breaks maintained
        ✅ Formula preservation
        ✅ No text corruption
        
        Approach: openpyxl for reading, xlsxwriter for writing
        """
        try:
            if not self.excel_path or not os.path.exists(self.excel_path):
                logger.error(f"Excel file not found: {self.excel_path}")
                return False
            
            # Step 1: Read existing data with openpyxl (preserves formulas)
            read_workbook = openpyxl.load_workbook(self.excel_path, rich_text=True)
            read_worksheet = read_workbook.active
            
            # Get all existing data including formulas AND formatting
            existing_data = []
            existing_formats = []
            
            for row_idx, row in enumerate(read_worksheet.iter_rows()):
                row_data = []
                row_formats = []
                for col_idx, cell in enumerate(row):
                    # Capture cell data
                    if cell.data_type == 'f':  # Formula
                        logger.info(f"Found formula at ({row_idx}, {col_idx}): {cell.value}")
                        row_data.append(('formula', cell.value))
                    elif hasattr(cell.value, '__class__') and cell.value.__class__.__name__ == 'CellRichText':
                        # Fix corrupted CellRichText objects from openpyxl reading xlsxwriter files
                        repaired_richtext = self._repair_corrupted_cellrichtext(cell.value)
                        row_data.append(('richtext', repaired_richtext))
                    else:
                        row_data.append(('value', cell.value))
                    
                    # Capture cell formatting with safe color extraction
                    def safe_color_extract(color_obj):
                        try:
                            if color_obj and hasattr(color_obj, 'rgb'):
                                return str(color_obj.rgb) if color_obj.rgb else None
                            return None
                        except:
                            return None
                    
                    cell_format = {
                        'font_bold': cell.font.bold if cell.font.bold else False,
                        'font_italic': cell.font.italic if cell.font.italic else False,
                        'font_color': safe_color_extract(cell.font.color),
                        'font_size': cell.font.size if cell.font.size else None,
                        'fill_color': safe_color_extract(cell.fill.start_color),
                        'alignment_wrap': cell.alignment.wrap_text if cell.alignment.wrap_text else False,
                        'alignment_horizontal': cell.alignment.horizontal if cell.alignment.horizontal else None,
                        'alignment_vertical': cell.alignment.vertical if cell.alignment.vertical else None,
                    }
                    row_formats.append(cell_format)
                
                existing_data.append(row_data)
                existing_formats.append(row_formats)
            
            # Capture column widths
            column_widths = {}
            for col_letter, dimension in read_worksheet.column_dimensions.items():
                if dimension.width:
                    column_widths[col_letter] = dimension.width
            
            # Capture row heights  
            row_heights = {}
            for row_num, dimension in read_worksheet.row_dimensions.items():
                if dimension.height:
                    row_heights[row_num] = dimension.height
            
            # Step 2: Create new file with xlsxwriter
            temp_file = f"{self.excel_path}.tmp"
            write_workbook = xlsxwriter.Workbook(temp_file)
            write_worksheet = write_workbook.add_worksheet()
            
            # Set default wrap text formatting for entire worksheet (A1:T100)
            wrap_format = write_workbook.add_format({'text_wrap': True})
            write_worksheet.set_column('A:T', None, wrap_format)
            
            # Apply column widths with wrap text formatting preserved
            for col_letter, width in column_widths.items():
                col_idx = openpyxl.utils.column_index_from_string(col_letter) - 1
                write_worksheet.set_column(col_idx, col_idx, width, wrap_format)
            
            # Apply row heights
            for row_num, height in row_heights.items():
                write_worksheet.set_row(row_num - 1, height)
            
            # Copy existing data with formatting to new workbook
            for row_idx, (row_data, row_formats) in enumerate(zip(existing_data, existing_formats)):
                for col_idx, ((data_type, value), cell_format) in enumerate(zip(row_data, row_formats)):
                    # Create xlsxwriter format from openpyxl format
                    format_dict = {}
                    if cell_format['font_bold']:
                        format_dict['bold'] = True
                    if cell_format['font_italic']:
                        format_dict['italic'] = True
                    if cell_format['font_color']:
                        # Convert color to valid xlsxwriter format
                        color_hex = self._convert_color_to_hex(cell_format['font_color'])
                        # Skip default black color to avoid overriding normal text
                        if color_hex and color_hex not in ['#000000', '#000']:
                            format_dict['color'] = color_hex
                    if cell_format['font_size']:
                        format_dict['size'] = cell_format['font_size']
                    if cell_format['fill_color']:
                        # Convert color to valid xlsxwriter format
                        color_hex = self._convert_color_to_hex(cell_format['fill_color'])
                        # Skip default white/transparent background to avoid overriding normal cells
                        if color_hex and color_hex not in ['#FFFFFF', '#FFF', '#000000', '#000']:
                            format_dict['bg_color'] = color_hex
                    # Always enable text wrap to preserve user's worksheet setting
                    format_dict['text_wrap'] = True
                    if cell_format['alignment_horizontal']:
                        format_dict['align'] = cell_format['alignment_horizontal']
                    if cell_format['alignment_vertical']:
                        format_dict['valign'] = cell_format['alignment_vertical']
                    
                    # Create format object - always include text_wrap to preserve worksheet setting
                    cell_format_obj = write_workbook.add_format(format_dict) if format_dict else write_workbook.add_format({'text_wrap': True})
                    
                    # Write data with formatting
                    if data_type == 'formula' and value:
                        logger.info(f"Writing formula at ({row_idx}, {col_idx}): {value}")
                        write_worksheet.write_formula(row_idx, col_idx, value, cell_format_obj)
                    elif data_type == 'richtext':
                        # Convert openpyxl RichText to xlsxwriter rich string
                        self._write_rich_text_xlsxwriter(write_worksheet, row_idx, col_idx, value, write_workbook)
                    elif value is not None:
                        write_worksheet.write(row_idx, col_idx, value, cell_format_obj)
                    elif cell_format_obj:
                        # Apply formatting even to empty cells
                        write_worksheet.write_blank(row_idx, col_idx, None, cell_format_obj)
            
            # Step 3: Add new row with rich text support  
            next_row = read_worksheet.max_row  # Use actual last row with data
            special_data = self._prepare_special_data(data, filename)
            
            # Create a default format with word wrap for new cells
            default_format = write_workbook.add_format({'text_wrap': True})
            
            for col_name, col_idx in self.columns.items():
                value = special_data.get(col_name, '')
                
                # Special handling for Dag column - create formula
                if col_name == 'Dag' and not value:
                    # Create formula =TEXT(I{row};"ddd") where I is the Tid start column
                    tid_start_col_idx = self.columns.get('Tid start', 9)  # Default to column I (9)
                    formula = f'=TEXT({get_column_letter(tid_start_col_idx)}{next_row + 1},"ddd")'
                    logger.info(f"Creating Dag formula for new row: {formula}")
                    write_worksheet.write_formula(next_row, col_idx-1, formula, default_format)
                elif hasattr(value, '__class__') and value.__class__.__name__ == 'CellRichText':
                    # Convert CellRichText to xlsxwriter rich string
                    self._write_rich_text_xlsxwriter(write_worksheet, next_row, col_idx-1, value, write_workbook)
                elif value is not None:
                    write_worksheet.write(next_row, col_idx-1, value, default_format)
                else:
                    # Apply word wrap format even to empty cells to maintain consistency
                    write_worksheet.write_blank(next_row, col_idx-1, None, default_format)
            
            write_workbook.close()
            
            # Step 4: Replace original file
            if os.path.exists(self.excel_path):
                os.replace(temp_file, self.excel_path)
            
            logger.info(f"Added row to Excel file using xlsxwriter hybrid approach at row {next_row + 1}")
            return True
            
        except Exception as e:
            logger.error(f"Error adding row with xlsxwriter: {e}")
            # Clean up temp file if it exists
            if 'temp_file' in locals() and os.path.exists(temp_file):
                os.remove(temp_file)
            return False
    
    def _prepare_special_data(self, data: Dict[str, str], filename: str) -> Dict[str, str]:
        """Prepare data with special column handling (extracted from original add_row)"""
        special_data = data.copy()
        
        # Händelse - preserve user content and add filename if needed
        if 'Händelse' in self.columns:
            current_content = data.get('Händelse', '')
            if hasattr(current_content, '__class__') and current_content.__class__.__name__ == 'CellRichText':
                has_content = len(current_content) > 0
                if has_content and filename:
                    special_data['Händelse'] = current_content
                elif not has_content:
                    special_data['Händelse'] = filename if filename else ""
                else:
                    special_data['Händelse'] = current_content
            else:
                current_content = str(current_content).strip()
                if current_content:
                    if filename and filename not in current_content:
                        special_data['Händelse'] = f"{current_content}\n{filename}"
                    else:
                        special_data['Händelse'] = current_content
                else:
                    if filename:
                        special_data['Händelse'] = f"\n\n{filename}"
                    else:
                        special_data['Händelse'] = ""
        
        # Tid start - use date from filename if user hasn't filled it in
        if 'Tid start' in self.columns and 'date' in data:
            user_tid_start = str(special_data.get('Tid start', '')).strip()
            if not user_tid_start:
                try:
                    date_obj = datetime.strptime(data['date'], '%Y-%m-%d')
                    special_data['Tid start'] = date_obj.date()
                except ValueError:
                    special_data['Tid start'] = data.get('date', '')
        
        # Källa1 - only use generated filename if field is empty AND we have a filename
        if 'Källa1' in self.columns:
            current_kalla1 = data.get('Källa1', '').strip()
            if not current_kalla1 and filename:
                special_data['Källa1'] = filename
            else:
                special_data['Källa1'] = current_kalla1
        
        return special_data
    
    def _convert_color_to_hex(self, color_value):
        """Convert various color formats to xlsxwriter-compatible hex string"""
        try:
            if not color_value:
                return None
            
            # Convert to string first
            color_str = str(color_value)
            
            # Remove any non-hex characters and get just the hex part
            import re
            hex_match = re.search(r'([0-9A-Fa-f]{6,8})', color_str)
            if hex_match:
                hex_color = hex_match.group(1)
                # Take last 6 characters (RGB) if 8 characters (ARGB)
                if len(hex_color) == 8:
                    hex_color = hex_color[2:]  # Remove alpha channel
                return f"#{hex_color}"
            
            # If no hex found, try some common color names
            color_lower = color_str.lower()
            color_map = {
                'black': '#000000',
                'white': '#FFFFFF',
                'red': '#FF0000',
                'green': '#00FF00',
                'blue': '#0000FF',
            }
            return color_map.get(color_lower, None)
            
        except Exception as e:
            logger.warning(f"Could not convert color {color_value}: {e}")
            return None
    
    def _write_rich_text_xlsxwriter(self, worksheet, row, col, rich_text_obj, workbook):
        """
        BREAKTHROUGH METHOD: Convert openpyxl CellRichText to xlsxwriter rich string
        
        TESTED CONFIRMED WORKING v1.7.4:
        ✅ All colors preserved (red, blue, green)
        ✅ Bold and italic formatting preserved
        ✅ Text order maintained correctly
        ✅ No text corruption or fragments
        """
        try:
            if not hasattr(rich_text_obj, '__iter__'):
                # Plain text
                worksheet.write(row, col, str(rich_text_obj))
                return
            
            # Build rich string for xlsxwriter
            rich_parts = []
            for part in rich_text_obj:
                if hasattr(part, 'text') and hasattr(part, 'font'):
                    # TextBlock with formatting
                    format_dict = {}
                    if hasattr(part.font, 'b') and part.font.b:
                        format_dict['bold'] = True
                    if hasattr(part.font, 'i') and part.font.i:
                        format_dict['italic'] = True
                    if hasattr(part.font, 'color') and part.font.color:
                        # Convert color to xlsxwriter format using the helper function
                        color_hex = self._convert_color_to_hex(part.font.color.rgb)
                        if color_hex and color_hex not in ['#000000', '#000']:
                            format_dict['color'] = color_hex
                    
                    if format_dict:
                        format_obj = workbook.add_format(format_dict)
                        rich_parts.extend([format_obj, part.text])
                    else:
                        rich_parts.append(part.text)
                elif isinstance(part, str):
                    # Plain text string part
                    rich_parts.append(part)
                else:
                    # Other type - convert to string
                    part_str = str(part)
                    rich_parts.append(part_str)
            
            # Use simple rich_string without complex filtering to avoid text ordering issues
            if rich_parts:
                worksheet.write_rich_string(row, col, *rich_parts)
            else:
                worksheet.write(row, col, str(rich_text_obj))
                
        except Exception as e:
            logger.warning(f"Error converting rich text to xlsxwriter format: {e}")
            # Fallback to plain text
            worksheet.write(row, col, str(rich_text_obj))

    def _repair_corrupted_cellrichtext(self, rich_text_obj):
        """
        Repair corrupted CellRichText objects that result from openpyxl reading xlsxwriter files.
        
        Problem: When xlsxwriter writes rich text and openpyxl reads it back, we get:
        Part 0: str('WHOLE TEXT CONTENT...')  # Duplicated full text
        Part 1: TextBlock('Formatted part 1')  # Individual formatted parts
        Part 2: TextBlock('Formatted part 2')  
        
        Solution: Remove the duplicated first part and keep only the TextBlocks.
        """
        try:
            if not hasattr(rich_text_obj, '__iter__') or len(rich_text_obj) <= 1:
                return rich_text_obj
                
            # Check if first part is a plain string containing most/all of the text
            first_part = rich_text_obj[0]
            if not isinstance(first_part, str):
                return rich_text_obj
                
            # Get total length of TextBlock parts
            textblock_length = 0
            textblock_count = 0
            for part in rich_text_obj[1:]:
                if hasattr(part, 'text') and hasattr(part, 'font'):  # TextBlock
                    textblock_length += len(part.text)
                    textblock_count += 1
                elif isinstance(part, str):
                    textblock_length += len(part)
                    
            # If first part is significantly longer than sum of other parts,
            # and we have TextBlocks, this is likely a corruption
            first_part_len = len(first_part)
            if (first_part_len > textblock_length * 0.7 and  # First part contains 70%+ of text
                textblock_count > 0):  # We have actual TextBlocks
                
                logger.info(f"REPAIR: Detected corrupted CellRichText - first part ({first_part_len} chars) vs TextBlocks ({textblock_length} chars)")
                
                # Create new CellRichText without the duplicated first part
                from openpyxl.cell.rich_text import CellRichText
                repaired_parts = list(rich_text_obj[1:])  # Skip first duplicated part
                
                if repaired_parts:
                    repaired = CellRichText(*repaired_parts)
                    logger.info(f"REPAIR: Created repaired CellRichText with {len(repaired)} parts")
                    return repaired
                else:
                    # Fallback to original if repair fails
                    logger.warning("REPAIR: No parts left after repair, using original")
                    return rich_text_obj
            else:
                # Not corrupted, return as-is
                return rich_text_obj
                
        except Exception as e:
            logger.warning(f"Error repairing CellRichText: {e}")
            return rich_text_obj


def get_formatted_text_for_excel_method2(text_widget):
    """
    METHOD 2: CHARACTER-BY-CHARACTER BREAKTHROUGH ALGORITHM
    
    TESTED CONFIRMED WORKING v1.7.4:
    ✅ Extracts all text content (no fragments)
    ✅ Preserves all formatting (colors, bold, italic)  
    ✅ Maintains line breaks and text order
    ✅ Creates proper CellRichText with TextBlocks
    
    This is the algorithm that replaced the broken tag-based approach
    and completely fixed the Excel rich text corruption issue.
    """
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        from openpyxl.styles.colors import Color
        
        # Get plain text
        plain_text = text_widget.get("1.0", "end-1c")
        
        # Check if there are any formatting tags
        all_tags = text_widget.tag_names()
        format_tags = [tag for tag in all_tags if tag in ["bold", "italic", "red", "blue", "green", "black"]]
        
        if not format_tags:
            # No formatting, return plain text
            return plain_text
        
        # METHOD 2: Process text character by character to maintain correct order
        rich_parts = []
        current_pos = "1.0"
        
        # Get all text
        text_end = text_widget.index("end-1c")
        
        # Iterate through text character by character
        while text_widget.compare(current_pos, "<", text_end):
            next_pos = f"{current_pos} +1c"
            char = text_widget.get(current_pos, next_pos)
            
            # Get tags at current position
            tags_at_pos = text_widget.tag_names(current_pos)
            format_tags_at_pos = [tag for tag in tags_at_pos if tag in ["bold", "italic", "red", "blue", "green", "black"]]
            
            # Build text with same formatting
            text_with_format = char
            temp_pos = next_pos
            
            # Continue until formatting changes
            while text_widget.compare(temp_pos, "<", text_end):
                temp_tags = text_widget.tag_names(temp_pos)
                temp_format_tags = [tag for tag in temp_tags if tag in ["bold", "italic", "red", "blue", "green", "black"]]
                
                if set(format_tags_at_pos) != set(temp_format_tags):
                    break
                    
                next_char_pos = f"{temp_pos} +1c"
                text_with_format += text_widget.get(temp_pos, next_char_pos)
                temp_pos = next_char_pos
            
            # Create appropriate part
            if format_tags_at_pos:
                font_kwargs = {}
                for tag in format_tags_at_pos:
                    if tag == "bold":
                        font_kwargs['b'] = True
                    elif tag == "italic":
                        font_kwargs['i'] = True
                    elif tag == "red":
                        font_kwargs['color'] = Color(rgb="FF0000")
                    elif tag == "blue":
                        font_kwargs['color'] = Color(rgb="0000FF")
                    elif tag == "green":
                        font_kwargs['color'] = Color(rgb="008000")
                    elif tag == "black":
                        font_kwargs['color'] = Color(rgb="000000")
                
                if font_kwargs:
                    font = InlineFont(**font_kwargs)
                    rich_parts.append(TextBlock(font, text_with_format))
                else:
                    rich_parts.append(text_with_format)
            else:
                rich_parts.append(text_with_format)
            
            current_pos = temp_pos
        
        # Create CellRichText if we have formatting
        if any(isinstance(part, TextBlock) for part in rich_parts):
            result = CellRichText(*rich_parts)
            return result
        else:
            return plain_text
            
    except Exception as e:
        # Fallback to plain text
        return text_widget.get("1.0", "end-1c")


# END OF BACKUP FILE
# This file contains all the working methods that make Excel rich text integration perfect in v1.7.4
# DO NOT MODIFY unless creating an improved version